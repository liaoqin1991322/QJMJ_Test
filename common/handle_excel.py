"""
coding:utf-8
@Software:Excel封装类，读写数据
@Time:2022/8/14 12:56
@Author:liaoqin
"""
import openpyxl


class HandleExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel_data(self):
        """
        从excel中读取数据"
        :return: 
        """""
        data_list = []
        # 1、加载 excel文件的工作簿对象
        workbook = openpyxl.load_workbook(self.file_name)
        # 2、选中表单
        sh = workbook[self.sheet_name]
        # 3、读取数据
        # 按行读取表单中所有的格子，每一行的格子放到一个元组中
        res = list(sh.rows)
        if len(res) > 0:
            # 取第一行
            rows_one = [i.value for i in res[0]]

            # 取剩余行数
            for v in res[1:]:
                datas = [i.value for i in v]
                # 组装成字典
                dic = dict(zip(rows_one, datas))
                data_list.append(dic)
        return data_list

    def write_excel_data(self, row, column, value):
        """
        数据写入excel的方法
        :param row:写入的行
        :param cloumn:写入的列
        :param value:写入的值
        :return:
        """
        pass
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        #写入数据
        sheet.cell(row=row,column=column,value=value)
        workbook.save(self.file_name)

# if __name__ == '__main__':
#     excel = HandleExcel(r'D:\pythonProject\QCD_Project\datas\test_case03.xlsx', '注册测试用例')
#     excel.write_excel_data(2,4,'------')

